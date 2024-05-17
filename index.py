# Imports from openpyxl
import pandas as pd
from openpyxl import load_workbook, Workbook # *** In python, imports are declared with the structure "from <module> import <class>"
from openpyxl.utils import get_column_letter


# define a main function
def main():
    # Define the paths
    # existing_file_path = 'data/input/test_input.xlsx'
    # existing_file_path = 'data/input/test_input_minimized.xlsx'
    existing_file_path = 'data/input/test_data_table.xlsx'
    new_file_path = 'data/output/test_output.xlsx'


    # Variables
    columns_to_keep = {
        'A':'Date',
        'C':'Description',
        'D':'Net Donation',
        'F':'Stripe Fee',
        'G':'Platform Fee',
        'H':'Total Gross Donation',
        'R':'Event',
        'X':'Source Title'
        }
    spreadsheet_title = None


    # Open the excel sheet with pandas
    df_donations = pd.read_excel(existing_file_path)
    
    # Sort the data by Description, then Event, then Source Title
    df_donations = df_donations.sort_values(by=['Source Title','Event','Description'])
    
    # Create a dictionary object to organize and store the sum of the revenue columns
    
    
    # Then remove first line, create a list of objects and sort through them
    # Analyze the data
        # perform logic
        # sum up the revenue columns
        # save the data
    # Use openpyxl to open the file
        # hide columns that aren't needed
        # append the title to the end of the sheet
        # save the file
    
    


    # openpyxl_operations(existing_file_path, new_file_path, columns_to_keep)







    # Save the workbook to a new file
    # print("Workbook saved to", new_file_path)





def openpyxl_operations(existing_file_path, new_file_path, columns_to_keep):
    
    # Load worksheet
    wb = load_workbook(filename=existing_file_path)
    ws = wb.active
    
    max_column = ws.max_column
    
    # 1. Get title and remove from first row
    if ws['A1'].value.startswith("Payout Report"): # *** In python, if statements are declared without parentheses and with a colon at the end, code block to be executed is indented instead of within brackets
        spreadsheet_title = ws['A1'].value
    else:
        print("Error: Spreadsheet title not found")
        # add other error handling code here
    print("Payout Report:", spreadsheet_title)

    ws.delete_rows(1)

    # 2. Hide columns except those that we want to keep

    # Convert to uppercase
    columns_to_keep = {k.upper():v for k,v in columns_to_keep.items()}
    # columns_to_keep = {k:v.upper() for k,v in columns_to_keep.items()}
    print("Columns to keep", columns_to_keep)

    # Loop through each column in the worksheet
    for col in range(1, max_column + 1):
        column_letter = get_column_letter(col)
        
        if column_letter not in columns_to_keep.keys():
            ws.column_dimensions[column_letter].hidden = True
        # Last. Append title to the end of sheet with buffer space
    ws.append([])
    ws.append([])
    ws.append([spreadsheet_title])
    
    # Save the workbook to a new file
    wb.save(new_file_path)

    # *** In python, instead of console logs, use "print(<message>)"
    print("Workbook saved to", new_file_path)



if __name__ == "__main__":
    print("\nRunning index.py")
    main()
    print("Finished running index.py\n")